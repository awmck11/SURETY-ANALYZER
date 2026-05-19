import pandas as pd
import yfinance as yf
from datetime import datetime
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# =====================================================
# CONFIGURATION
# =====================================================

VOLATILITY_THRESHOLD = 0.75  # 75% spread vs. mean = volatile

# Color codes for flag formatting
RED_FILL = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")
GREEN_FILL = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SECTION_FONT = Font(bold=True, size=11)
BOLD_FONT = Font(bold=True)


# =====================================================
# RATIO CALCULATIONS
# =====================================================

def calc_wc(ca, cl): return ca - cl
def calc_cr(ca, cl): return ca / cl
def calc_qr(ca, inv, pp, cl): return (ca - inv - pp) / cl
def calc_de(td, te): return td / te
def calc_er(te, ta): return te / ta
def calc_gpm(gp, rev): return gp / rev
def calc_npm(ni, rev): return ni / rev
def calc_roe(ni, te): return ni / te
def calc_wc_rev(wc, rev): return wc / rev


# =====================================================
# SNAPSHOT FLAGGING
# =====================================================

def flag_current_ratio(r):
    if r < 1.0: return "RED FLAG"
    elif r < 1.2: return "TIGHT"
    elif r < 1.5: return "ACCEPTABLE"
    elif r < 2.0: return "STRONG"
    else: return "VERY STRONG"

def flag_quick_ratio(r):
    if r < 0.8: return "RED FLAG"
    elif r < 1.0: return "TIGHT"
    elif r < 1.5: return "ACCEPTABLE"
    else: return "STRONG"

def flag_de(r):
    if r > 3.0: return "RED FLAG"
    elif r > 2.0: return "HIGH RISK"
    elif r > 1.0: return "MODERATE"
    elif r > 0.5: return "CONSERVATIVE"
    else: return "VERY CONSERVATIVE"

def flag_er(r):
    if r < 0.15: return "RED FLAG"
    elif r < 0.25: return "TIGHT"
    elif r < 0.40: return "ACCEPTABLE"
    else: return "STRONG"

def flag_gpm_v(m):
    if m < 0.05: return "RED FLAG"
    elif m < 0.10: return "TIGHT"
    elif m < 0.20: return "ACCEPTABLE"
    else: return "STRONG"

def flag_npm_v(m):
    if m < 0: return "RED FLAG"
    elif m < 0.02: return "TIGHT"
    elif m < 0.05: return "ACCEPTABLE"
    else: return "STRONG"

def flag_roe_v(r):
    if r < 0: return "RED FLAG"
    elif r < 0.05: return "WEAK"
    elif r < 0.15: return "ACCEPTABLE"
    else: return "STRONG"

def flag_wc_rev_v(r):
    if r < 0.05: return "RED FLAG"
    elif r < 0.08: return "TIGHT"
    elif r < 0.15: return "ACCEPTABLE"
    else: return "STRONG"


# =====================================================
# TREND FLAGGING
# =====================================================

def flag_trend(values, higher_is_better=True):
    clean = [v for v in values if v is not None and not pd.isna(v)]
    if len(clean) < 2:
        return "INSUFFICIENT DATA"

    first = clean[0]
    last = clean[-1]
    change = (last - first) / abs(first) if first != 0 else 0
    avg = sum(clean) / len(clean)
    spread = (max(clean) - min(clean)) / abs(avg) if avg != 0 else 0

    if spread > VOLATILITY_THRESHOLD:
        return "VOLATILE"

    threshold = 0.10
    if higher_is_better:
        if change > threshold: return "IMPROVING"
        elif change < -threshold: return "DETERIORATING"
        else: return "STABLE"
    else:
        if change > threshold: return "DETERIORATING"
        elif change < -threshold: return "IMPROVING"
        else: return "STABLE"


def calc_growth_rate(values):
    clean = [v for v in values if v is not None and not pd.isna(v)]
    if len(clean) < 2: return None
    growth_rates = []
    for i in range(1, len(clean)):
        if clean[i-1] != 0:
            growth_rates.append((clean[i] - clean[i-1]) / abs(clean[i-1]))
    return sum(growth_rates) / len(growth_rates) if growth_rates else None


def flag_growth(g):
    if g is None: return "INSUFFICIENT DATA"
    if g > 0.40: return "RED FLAG - rapid growth"
    elif g > 0.20: return "ELEVATED"
    elif g > 0.05: return "HEALTHY GROWTH"
    elif g > -0.05: return "FLAT"
    else: return "DECLINING"


# =====================================================
# OVERALL UNDERWRITING VERDICT (Moderate personality)
# =====================================================

def compute_verdict(snapshot, trends):
    """Counts red flags and tight flags, applies moderate underwriting logic."""
    red_count = 0
    tight_count = 0
    strong_count = 0

    for flag in snapshot.values():
        if "RED FLAG" in flag:
            red_count += 1
        elif flag in ["TIGHT", "WEAK"]:
            tight_count += 1
        elif "STRONG" in flag:
            strong_count += 1

    # Trend adjustments
    trend_warnings = 0
    if trends.get("Gross Margin Trend") == "DETERIORATING": trend_warnings += 1
    if trends.get("Net Margin Trend") == "DETERIORATING": trend_warnings += 1
    if trends.get("D/E Trend") == "DETERIORATING": trend_warnings += 1
    if trends.get("Revenue Growth Flag") == "RED FLAG - rapid growth": trend_warnings += 1
    if trends.get("Equity Growth Flag") == "DECLINING": trend_warnings += 1

    # Moderate underwriting logic
    if red_count >= 3 or trend_warnings >= 3:
        verdict = "DECLINE"
        reasoning = "Multiple red flags or deteriorating trends across pillars."
    elif red_count >= 2 or (red_count >= 1 and trend_warnings >= 2):
        verdict = "DECLINE OR HEAVILY CONDITIONED"
        reasoning = "Significant credit concerns; bondable only with collateral or capacity restrictions."
    elif red_count == 1 or tight_count >= 3 or trend_warnings >= 2:
        verdict = "APPROVE WITH CONDITIONS"
        reasoning = "Bondable with monitoring; specific items require explanation or follow-up."
    elif tight_count >= 1 or trend_warnings >= 1:
        verdict = "APPROVE - STANDARD MONITORING"
        reasoning = "Acceptable risk with routine quarterly review."
    else:
        verdict = "APPROVE - PREFERRED"
        reasoning = "Strong across all pillars; candidate for expanded capacity."

    return verdict, reasoning, red_count, tight_count, strong_count, trend_warnings


# =====================================================
# DATA FETCHING
# =====================================================

def safe_get(df, row_name, column):
    try:
        value = df.loc[row_name, column]
        if pd.isna(value): return None
        return float(value)
    except (KeyError, IndexError):
        return None


def fetch_multi_year_data(ticker):
    print(f"Fetching multi-year data for {ticker}...")
    company = yf.Ticker(ticker)
    info = company.info
    name = info.get("longName", ticker)

    bs = company.balance_sheet
    is_ = company.income_stmt

    bs_columns = list(bs.columns)[::-1]
    is_columns = list(is_.columns)[::-1]

    years_data = []
    for bs_col, is_col in zip(bs_columns, is_columns):
        years_data.append({
            "year": bs_col.year,
            "current_assets": safe_get(bs, "Current Assets", bs_col),
            "current_liabilities": safe_get(bs, "Current Liabilities", bs_col),
            "inventory": safe_get(bs, "Inventory", bs_col) or 0,
            "prepaids": safe_get(bs, "Other Current Assets", bs_col) or 0,
            "total_assets": safe_get(bs, "Total Assets", bs_col),
            "total_equity": safe_get(bs, "Stockholders Equity", bs_col),
            "total_debt": safe_get(bs, "Total Debt", bs_col),
            "revenue": safe_get(is_, "Total Revenue", is_col),
            "gross_profit": safe_get(is_, "Gross Profit", is_col),
            "net_income": safe_get(is_, "Net Income", is_col),
        })

    return {"name": name, "ticker": ticker, "years": years_data}


# =====================================================
# ANALYSIS
# =====================================================

def analyze(contractor):
    name = contractor["name"]
    ticker = contractor["ticker"]
    years = contractor["years"]

    rows = []
    revs, eqs, gpms, npms, des = [], [], [], [], []

    for y in years:
        if not y["current_assets"] or not y["current_liabilities"]:
            continue
        wc = calc_wc(y["current_assets"], y["current_liabilities"])
        cr = calc_cr(y["current_assets"], y["current_liabilities"])
        qr = calc_qr(y["current_assets"], y["inventory"], y["prepaids"], y["current_liabilities"])
        de = calc_de(y["total_debt"], y["total_equity"]) if y["total_debt"] and y["total_equity"] else None
        er = calc_er(y["total_equity"], y["total_assets"]) if y["total_equity"] and y["total_assets"] else None
        gpm = calc_gpm(y["gross_profit"], y["revenue"]) if y["gross_profit"] and y["revenue"] else None
        npm = calc_npm(y["net_income"], y["revenue"]) if y["net_income"] and y["revenue"] else None
        roe = calc_roe(y["net_income"], y["total_equity"]) if y["net_income"] and y["total_equity"] else None
        wc_rev = calc_wc_rev(wc, y["revenue"]) if y["revenue"] else None

        rows.append({
            "Year": y["year"],
            "Revenue ($M)": y["revenue"] / 1e6 if y["revenue"] else None,
            "Equity ($M)": y["total_equity"] / 1e6 if y["total_equity"] else None,
            "Working Capital ($M)": wc / 1e6,
            "Current Ratio": cr,
            "Quick Ratio": qr,
            "Debt-to-Equity": de,
            "Equity Ratio": er,
            "Gross Profit Margin": gpm,
            "Net Profit Margin": npm,
            "Return on Equity": roe,
            "WC to Revenue": wc_rev,
        })

        if y["revenue"]: revs.append(y["revenue"])
        if y["total_equity"]: eqs.append(y["total_equity"])
        if gpm is not None: gpms.append(gpm)
        if npm is not None: npms.append(npm)
        if de is not None: des.append(de)

    trends = {
        "Revenue Growth": calc_growth_rate(revs),
        "Revenue Growth Flag": flag_growth(calc_growth_rate(revs)),
        "Equity Growth": calc_growth_rate(eqs),
        "Equity Growth Flag": flag_growth(calc_growth_rate(eqs)),
        "Gross Margin Trend": flag_trend(gpms, higher_is_better=True),
        "Net Margin Trend": flag_trend(npms, higher_is_better=True),
        "D/E Trend": flag_trend(des, higher_is_better=False),
    }

    snapshot = {}
    latest = rows[-1] if rows else None
    if latest:
        snapshot = {
            "Current Ratio": flag_current_ratio(latest["Current Ratio"]),
            "Quick Ratio": flag_quick_ratio(latest["Quick Ratio"]),
            "Debt-to-Equity": flag_de(latest["Debt-to-Equity"]) if latest["Debt-to-Equity"] else "N/A",
            "Equity Ratio": flag_er(latest["Equity Ratio"]) if latest["Equity Ratio"] else "N/A",
            "Gross Profit Margin": flag_gpm_v(latest["Gross Profit Margin"]) if latest["Gross Profit Margin"] else "N/A",
            "Net Profit Margin": flag_npm_v(latest["Net Profit Margin"]) if latest["Net Profit Margin"] else "N/A",
            "Return on Equity": flag_roe_v(latest["Return on Equity"]) if latest["Return on Equity"] else "N/A",
            "WC to Revenue": flag_wc_rev_v(latest["WC to Revenue"]) if latest["WC to Revenue"] else "N/A",
        }

    verdict, reasoning, reds, tights, strongs, trend_warnings = compute_verdict(snapshot, trends)

    return {
        "name": name, "ticker": ticker, "rows": rows,
        "trends": trends, "snapshot": snapshot,
        "verdict": verdict, "reasoning": reasoning,
        "red_count": reds, "tight_count": tights,
        "strong_count": strongs, "trend_warnings": trend_warnings,
    }


# =====================================================
# CONSOLE OUTPUT
# =====================================================

def print_analysis(a):
    print("=" * 80)
    print(f"  {a['name']} ({a['ticker']})")
    print("=" * 80)
    print(f"  UNDERWRITING VERDICT: {a['verdict']}")
    print(f"  {a['reasoning']}")
    print(f"  Red Flags: {a['red_count']} | Tight: {a['tight_count']} | Strong: {a['strong_count']} | Trend Warnings: {a['trend_warnings']}")
    print()
    print("  SCHLEIFER TREND ANALYSIS")
    rg = a['trends']['Revenue Growth']
    eg = a['trends']['Equity Growth']
    print(f"    Revenue Growth:     {rg:.2%}  -> {a['trends']['Revenue Growth Flag']}" if rg else "    Revenue Growth: N/A")
    print(f"    Equity Growth:      {eg:.2%}  -> {a['trends']['Equity Growth Flag']}" if eg else "    Equity Growth: N/A")
    print(f"    Gross Margin Trend: {a['trends']['Gross Margin Trend']}")
    print(f"    Net Margin Trend:   {a['trends']['Net Margin Trend']}")
    print(f"    D/E Trend:          {a['trends']['D/E Trend']}")
    print()


# =====================================================
# EXCEL EXPORT (formatted)
# =====================================================

def get_fill_for_flag(flag):
    """Returns the appropriate fill color for a flag string."""
    if not flag or flag == "N/A":
        return None
    flag_upper = str(flag).upper()
    if "RED FLAG" in flag_upper or "HIGH RISK" in flag_upper or "DETERIORATING" in flag_upper or "DECLINING" in flag_upper or "DECLINE" in flag_upper:
        return RED_FILL
    if "TIGHT" in flag_upper or "WEAK" in flag_upper or "MODERATE" in flag_upper or "ELEVATED" in flag_upper or "VOLATILE" in flag_upper or "CONDITIONS" in flag_upper:
        return YELLOW_FILL
    if "STRONG" in flag_upper or "IMPROVING" in flag_upper or "HEALTHY" in flag_upper or "ACCEPTABLE" in flag_upper or "CONSERVATIVE" in flag_upper or "PREFERRED" in flag_upper or "APPROVE" in flag_upper:
        return GREEN_FILL
    return None


def export_excel(analyses, filename):
    with pd.ExcelWriter(filename, engine="openpyxl") as writer:

        # ============ EXECUTIVE SUMMARY TAB ============
        summary_data = []
        for a in analyses:
            summary_data.append({
                "Company": a["name"],
                "Ticker": a["ticker"],
                "Verdict": a["verdict"],
                "Reasoning": a["reasoning"],
                "Red Flags": a["red_count"],
                "Tight Flags": a["tight_count"],
                "Strong Metrics": a["strong_count"],
                "Trend Warnings": a["trend_warnings"],
                "Revenue Growth": a["trends"]["Revenue Growth"],
                "Equity Growth": a["trends"]["Equity Growth"],
                "GPM Trend": a["trends"]["Gross Margin Trend"],
                "NPM Trend": a["trends"]["Net Margin Trend"],
                "D/E Trend": a["trends"]["D/E Trend"],
            })
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name="Executive Summary", index=False)

        ws = writer.sheets["Executive Summary"]

        # Format header row
        for cell in ws[1]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Color the verdict, trend, and growth-flag cells
        for row_idx, a in enumerate(analyses, start=2):
            # Verdict column (C)
            ws.cell(row=row_idx, column=3).fill = get_fill_for_flag(a["verdict"]) or PatternFill()
            ws.cell(row=row_idx, column=3).font = BOLD_FONT
            # GPM Trend (K), NPM Trend (L), D/E Trend (M)
            ws.cell(row=row_idx, column=11).fill = get_fill_for_flag(a["trends"]["Gross Margin Trend"]) or PatternFill()
            ws.cell(row=row_idx, column=12).fill = get_fill_for_flag(a["trends"]["Net Margin Trend"]) or PatternFill()
            ws.cell(row=row_idx, column=13).fill = get_fill_for_flag(a["trends"]["D/E Trend"]) or PatternFill()
            # Format growth as percentage
            ws.cell(row=row_idx, column=9).number_format = "0.00%"
            ws.cell(row=row_idx, column=10).number_format = "0.00%"

        # Column widths
        widths = [32, 8, 32, 50, 10, 10, 14, 14, 14, 14, 16, 16, 16]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        ws.row_dimensions[1].height = 30
        ws.freeze_panes = "A2"

        # ============ PER-CONTRACTOR TABS ============
        for a in analyses:
            sheet_name = a["ticker"][:30]
            df = pd.DataFrame(a["rows"])
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            ws = writer.sheets[sheet_name]

            # Header formatting
            for cell in ws[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            num_rows = len(a["rows"]) + 1  # +1 for header

            # Number formatting on data rows
            for row in range(2, num_rows + 1):
                ws.cell(row=row, column=2).number_format = "#,##0"  # Revenue
                ws.cell(row=row, column=3).number_format = "#,##0"  # Equity
                ws.cell(row=row, column=4).number_format = "#,##0"  # WC
                ws.cell(row=row, column=5).number_format = "0.00"   # CR
                ws.cell(row=row, column=6).number_format = "0.00"   # QR
                ws.cell(row=row, column=7).number_format = "0.00"   # D/E
                ws.cell(row=row, column=8).number_format = "0.00%"  # Equity Ratio
                ws.cell(row=row, column=9).number_format = "0.00%"  # GPM
                ws.cell(row=row, column=10).number_format = "0.00%" # NPM
                ws.cell(row=row, column=11).number_format = "0.00%" # ROE
                ws.cell(row=row, column=12).number_format = "0.00%" # WC/Rev

            # Add a TREND ANALYSIS section below the year data
            section_start = num_rows + 2
            ws.cell(row=section_start, column=1).value = "SCHLEIFER TREND ANALYSIS"
            ws.cell(row=section_start, column=1).font = SECTION_FONT
            ws.cell(row=section_start, column=1).fill = SECTION_FILL

            trend_items = [
                ("Revenue Growth (avg YoY)", a["trends"]["Revenue Growth"], a["trends"]["Revenue Growth Flag"], "0.00%"),
                ("Equity Growth (avg YoY)", a["trends"]["Equity Growth"], a["trends"]["Equity Growth Flag"], "0.00%"),
                ("Gross Margin Trend", None, a["trends"]["Gross Margin Trend"], None),
                ("Net Margin Trend", None, a["trends"]["Net Margin Trend"], None),
                ("D/E Trend", None, a["trends"]["D/E Trend"], None),
            ]
            for i, (label, value, flag, fmt) in enumerate(trend_items, start=1):
                r = section_start + i
                ws.cell(row=r, column=1).value = label
                ws.cell(row=r, column=1).font = BOLD_FONT
                if value is not None:
                    ws.cell(row=r, column=2).value = value
                    ws.cell(row=r, column=2).number_format = fmt
                ws.cell(row=r, column=3).value = flag
                fill = get_fill_for_flag(flag)
                if fill:
                    ws.cell(row=r, column=3).fill = fill

            # Latest year snapshot section
            snapshot_start = section_start + len(trend_items) + 3
            ws.cell(row=snapshot_start, column=1).value = "LATEST YEAR SNAPSHOT"
            ws.cell(row=snapshot_start, column=1).font = SECTION_FONT
            ws.cell(row=snapshot_start, column=1).fill = SECTION_FILL

            for i, (metric, flag) in enumerate(a["snapshot"].items(), start=1):
                r = snapshot_start + i
                ws.cell(row=r, column=1).value = metric
                ws.cell(row=r, column=1).font = BOLD_FONT
                ws.cell(row=r, column=3).value = flag
                fill = get_fill_for_flag(flag)
                if fill:
                    ws.cell(row=r, column=3).fill = fill

            # Verdict section
            verdict_start = snapshot_start + len(a["snapshot"]) + 3
            ws.cell(row=verdict_start, column=1).value = "UNDERWRITING VERDICT"
            ws.cell(row=verdict_start, column=1).font = SECTION_FONT
            ws.cell(row=verdict_start, column=1).fill = SECTION_FILL

            ws.cell(row=verdict_start + 1, column=1).value = "Verdict"
            ws.cell(row=verdict_start + 1, column=1).font = BOLD_FONT
            ws.cell(row=verdict_start + 1, column=3).value = a["verdict"]
            fill = get_fill_for_flag(a["verdict"])
            if fill:
                ws.cell(row=verdict_start + 1, column=3).fill = fill
            ws.cell(row=verdict_start + 1, column=3).font = BOLD_FONT

            ws.cell(row=verdict_start + 2, column=1).value = "Reasoning"
            ws.cell(row=verdict_start + 2, column=1).font = BOLD_FONT
            ws.cell(row=verdict_start + 2, column=3).value = a["reasoning"]
            ws.cell(row=verdict_start + 2, column=3).alignment = Alignment(wrap_text=True)

            # Column widths
            ws.column_dimensions["A"].width = 32
            for col in range(2, 13):
                ws.column_dimensions[get_column_letter(col)].width = 14
            ws.column_dimensions["C"].width = 45  # For flag text in trend/snapshot sections

            ws.row_dimensions[1].height = 30
            ws.freeze_panes = "A2"

    print(f"Excel report saved to: {filename}")


# =====================================================
# MAIN
# =====================================================

tickers = ["GVA", "TPC", "STRL"]

analyses = []
for ticker in tickers:
    contractor = fetch_multi_year_data(ticker)
    analysis = analyze(contractor)
    print_analysis(analysis)
    analyses.append(analysis)

timestamp = datetime.now().strftime("%Y%m%d_%H%M")
filename = f"surety_analysis_{timestamp}.xlsx"
export_excel(analyses, filename)